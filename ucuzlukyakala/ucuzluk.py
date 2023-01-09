from bs4 import BeautifulSoup as bs
import requests
import datetime
from datetime import date
from urllib.request import urlopen
import xml.etree.ElementTree as et
import pandas as pd
import matplotlib.pyplot as plt
#import datetime
#import sqlite3
import openpyxl







url="https://www.cimri.com/market/toz-ve-kup-seker/en-ucuz-balkupu-5-kg-toz-seker-fiyatlari,51432?cimriRedirect=true"
page=requests.get(url).content
soup = bs(page, 'html.parser')
balkupu=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/surulebilir/en-ucuz-abdurrahman-tatlici-tadibu-330-gr-kakaolu-findik-ezmesi-fiyatlari,19638"
page=requests.get(url).content
soup = bs(page, 'html.parser')
tadıbu=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/parlatici-ve-temizleyiciler/en-ucuz-domestos-aninda-hijyen-200-ml-pratik-yuzey-spreyi-fiyatlari,78235?cimriRedirect=true"
page=requests.get(url).content
soup = bs(page, 'html.parser')
spray=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/un-ve-irmik/en-ucuz-city-farm-1-kg-organik-tam-bugday-unu-fiyatlari,36656"
page=requests.get(url).content
soup = bs(page, 'html.parser')
tambugday=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/un-ve-irmik/en-ucuz-soke-5-kg-geleneksel-un-fiyatlari,36625"
page=requests.get(url).content
soup = bs(page, 'html.parser')
beyazun=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/kuru-kedi-mamasi/en-ucuz-nd-dusuk-tahilli-somonlu-kisirlastirilmis-10-kg-kedi-mamasi-fiyatlari,483057920"
page=requests.get(url).content
soup = bs(page, 'html.parser')
kedi=(soup.find('span',{"class":"s1wl91l5-4 cBVHJG"}))
kedilink=soup.find("a",{"class":"notOnlyApp s1wl91l5-0 cZnMAi"})

url="https://www.cimri.com/bitkisel-besin-destekleri/en-ucuz-gifrer-40li-serum-fizyolojik-damla-fiyatlari,620263430"
page=requests.get(url).content
soup = bs(page, 'html.parser')
serumfizyo=(soup.find('span',{"class":"s1wl91l5-4 cBVHJG"}))

url="https://www.cimri.com/pratik-mutfak-gerecleri/en-ucuz-brita-maxtra-yedek-su-filtresi-altili-fiyatlari,782728844"
page=requests.get(url).content
soup = bs(page, 'html.parser')
brita=(soup.find('span',{"class":"s1wl91l5-4 cBVHJG"}))

url="https://www.cimri.com/market/kahvaltilik-gevrek/en-ucuz-eti-lifalif-500-gr-yulaf-ezmesi-fiyatlari,50409"
page=requests.get(url).content
soup = bs(page, 'html.parser')
yulaf=(soup.find('span',{"class":"OfferCard_price__en-Ud"}))

url="https://www.cimri.com/ofis-sandalyesi/en-ucuz-xdrive-15li-profesyonel-oyuncu-koltugu-kirmizisiyah-xdrive-fiyatlari,695237781"
page=requests.get(url).content
soup = bs(page, 'html.parser')
xdrive=(soup.find('span',{"class":"s1wl91l5-4 cBVHJG"}))

url="https://www.cimri.com/market/peynir/en-ucuz-sutas-600-gr-kasar-peyniri-fiyatlari,16718?cimriRedirect=true"
page=requests.get(url).content
soup = bs(page, 'html.parser')
kasar=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/unlu-mamuller/en-ucuz-uno-550-gr-buyuk-tost-ekmegi-fiyatlari,52394"
page=requests.get(url).content
soup = bs(page, 'html.parser')
uno=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/sut/en-ucuz-pinar-1-lt-sut-fiyatlari,17206"
page=requests.get(url).content
soup = bs(page, 'html.parser')
sut=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/gazli-icecek/en-ucuz-coca-cola-2-5-lt-fiyatlari,63334"
page=requests.get(url).content
soup = bs(page, 'html.parser')
kola=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/cay/en-ucuz-caykur-filiz-1-kg-cay-fiyatlari,54368"
page=requests.get(url).content
soup = bs(page, 'html.parser')
cay=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/sivi-yag-ve-zeytinyagi/en-ucuz-yudum-5-lt-teneke-aycicek-yagi-fiyatlari,7"
page=requests.get(url).content
soup = bs(page, 'html.parser')
aycicek=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.kuruyemisonline.com/urun/findik-ici-kavrulmus-250gr"
page=requests.get(url).content
soup = bs(page, 'html.parser')
findik=(soup.find('div',{"class":"product-price-old"}))

url="https://www.cimri.com/market/makarnalar/en-ucuz-barilla-fusilli-500-gr-burgu-makarna-fiyatlari,42611"
page=requests.get(url).content
soup = bs(page, 'html.parser')
makarna=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))

url="https://www.cimri.com/market/bakliyat/en-ucuz-yayla-2-kg-osmancik-pirinc-fiyatlari,20365"
page=requests.get(url).content
soup = bs(page, 'html.parser')
pirinc=(soup.find('div',{"class":"MainOfferCard_price_container__22jHp"}))







urunler=["Balküpü 5 kg toz şeker","Abdurrahman Tatlıcı Tadıbu","Domestos Spray 200 ml","City farm tam buğday 1 kg",
         "Söke Un 5 kg","N&D Kısırlaştırılmış kedi maması","Gifrer Serum Fizyolojik 40'lı","Brita Maxtra 6'lı filtre",
         "Eti Lifalif Yulaf","X Drive 15'li","Sütaş 600 gram Kaşar Peyniri","Uno 550 gram Büyük Tost Ekmeği",
         "Pınar 1 Litre Tam Yağlı Süt","Coca Cola 2.5 Litre","Çaykur Filiz 1 kg Çay","Yudum 5 lt Teneke Ayçiçek Yağı","250 gram Fındık","Barilla 500 gram Makarna",
         "Yayla 2 kg Osmancık Pirinç"]
fiyatlar=[balkupu.text,tadıbu.text,spray.text,tambugday.text,beyazun.text,kedi.text,serumfizyo.text,brita.text,yulaf.text,
          xdrive.text,kasar.text,uno.text,sut.text,kola.text,cay.text,aycicek.text,findik.text,makarna.text,pirinc.text]
column=date.today()
sontablo=pd.DataFrame({str(column):fiyatlar},index=urunler)

def newcol(sheet_name, column):

    ws = wb[sheet_name]
    new_column = ws.max_column + 1

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)



wb=openpyxl.load_workbook("ucuzlukyakalatablo.xlsx")
newcol("Sheet1",[str(column),balkupu.text,tadıbu.text,spray.text,tambugday.text,beyazun.text,kedi.text,serumfizyo.text,brita.text,yulaf.text,
                 xdrive.text,kasar.text,uno.text,sut.text,kola.text,cay.text,aycicek.text,findik.text,makarna.text,pirinc.text])
wb.save("ucuzlukyakalatablo.xlsx")
print("Update başarılı!")



'''
df=pd.read_excel("ucuzlukyakalatablo.xlsx")
df.rename(columns = {'Unnamed: 0':'Urunler'}, inplace = True)
df2=df.set_index("Urunler")


writer = pd.ExcelWriter('ucuzlukyakalatablo.xlsx')
sontablo.to_excel(writer)
writer.save()
print("DataFrame is exported successfully to 'ucuzlukyakalatablo.xlsx' Excel File.")
'''





